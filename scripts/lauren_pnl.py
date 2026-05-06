"""
lauren_pnl — read Lauren's master P&L Google Sheet directly.

The Sheet is publicly readable (anyone with link can view), so we
fetch it as XLSX via the public export URL — no auth needed, no
Service Account setup, no API key, no Compliance review. Works
from GitHub Actions cloud runners and from local Cowork sessions.

Usage:

    from scripts.lauren_pnl import fetch_workbook, get_event_data

    wb = fetch_workbook()
    data = get_event_data(wb, "Nashville")
    # data is a dict with: net_sales, num_tx, avg_ticket, profit,
    # inventory[], expenses[], qb_transactions[], ...

The Sheet ID is hard-coded as a constant. If Lauren ever swaps the
master Sheet, change SHEET_ID here.

Authored 2026-05-05 after Lauren said her preferred QB workflow is
to paste transactions into the Sheet (her existing habit) rather
than upload separate CSVs to the modal.
"""

import datetime as _dt
import io as _io
import urllib.request as _urlreq
from pathlib import Path as _Path

try:
    import openpyxl  # workbook parser
except ImportError as _e:
    raise SystemExit(
        "openpyxl is required. In a workflow YAML add a step:\n"
        "    pip install --quiet openpyxl"
    ) from _e


# ---------------------------------------------------------------------------
# Constants — Lauren's master P&L Sheet
# ---------------------------------------------------------------------------

SHEET_ID = "1mUn3foJhkJCje9LmJVErxKfFwSWJDgcFaZkg_7oWCM4"
EXPORT_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"

# Column map — discovered via inspection on 2026-05-05.
# All columns are 1-indexed (col A = 1).
COL_DATE       = 1   # A — start date of event (only in row 2)
COL_LOCATION   = 2   # B — city name (row 2). ALSO inventory item names rows 5+.
COL_NET_SALES  = 3   # C — total Net Sales (row 2)
COL_NUM_TX     = 4   # D — # transactions (row 2)
COL_AVG_TICKET = 5   # E — avg sale (row 2)
COL_PROFIT     = 6   # F — final profit cell (row 2)

COL_INV_NAME   = 1   # A — inventory item names (rows 5+)
COL_INV_AMT    = 2   # B — inventory amounts (rows 5+)
COL_MONTHLY_LBL = 4  # D — Monthly expenses labels (Miscellaneous / Gas)
COL_MONTHLY_AMT = 5  # E — Monthly expenses amounts

COL_EXP_LABEL  = 8   # H — expense category labels (rows 2+)
COL_EXP_AMT    = 9   # I — expense category amounts (rows 2+)

# The QB paste area is NOT at a fixed column — different tabs anchor it
# at K, L, sometimes M. We detect it dynamically by finding the cell in
# row 1 whose value is "DATE" (that's the QB-area date header).
QB_HEADERS = ["DATE", "From/To", "Category", "MEMO", "TYPE", "Amount", "Tags"]


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def fetch_workbook(timeout: int = 30):
    """Pull the latest XLSX of Lauren's P&L Sheet and parse with openpyxl.

    Returns an `openpyxl.Workbook`. The workbook is loaded with `data_only=True`
    so cell values reflect Google Sheets' last cached formula result (not
    the raw formulas). That's what we want for reading P&L numbers.
    """
    req = _urlreq.Request(EXPORT_URL, headers={"User-Agent": "lauren-pnl/1.0"})
    with _urlreq.urlopen(req, timeout=timeout) as r:
        raw = r.read()
    return openpyxl.load_workbook(_io.BytesIO(raw), data_only=True)


def list_event_tabs(wb) -> list:
    """Return the list of event tab names in the workbook (in order)."""
    return list(wb.sheetnames)


def get_event_data(wb, tab_name: str) -> dict:
    """Extract every relevant number from one event's tab.

    `tab_name` is the city name as it appears in the workbook
    (e.g. "Nashville", "Glendale", "Fort Worth"). Match is case-sensitive.

    Returns a dict:
        {
          "tab":            "Nashville",
          "start_date":     date or None,
          "city":           "Nashville",
          "net_sales":      float or None,
          "num_tx":         int or None,
          "avg_ticket":     float or None,
          "profit":         float or None,
          "inventory":      [{"name": "She", "amount": -2988.60}, ...],
          "monthly_expenses": [{"name": "Miscellaneous", "amount": -6000}, ...],
          "expenses":       [{"name": "Goni", "amount": -1120}, ...],
          "qb_transactions": [
            {"date": ..., "from_to": "...", "category": "...",
             "memo": "...", "type": "...", "amount": ..., "tags": "..."},
            ...
          ],
          "is_complete":    bool      # True if profit cell is non-empty
        }
    """
    if tab_name not in wb.sheetnames:
        raise KeyError(
            f"Tab {tab_name!r} not found. Available: {wb.sheetnames}"
        )
    sh = wb[tab_name]

    out = {
        "tab":              tab_name,
        "start_date":       sh.cell(row=2, column=COL_DATE).value,
        "city":             sh.cell(row=2, column=COL_LOCATION).value,
        "net_sales":        _num(sh.cell(row=2, column=COL_NET_SALES).value),
        "num_tx":           _int(sh.cell(row=2, column=COL_NUM_TX).value),
        "avg_ticket":       _num(sh.cell(row=2, column=COL_AVG_TICKET).value),
        "profit":           _num(sh.cell(row=2, column=COL_PROFIT).value),
        "inventory":        [],
        "monthly_expenses": [],
        "expenses":         [],
        "qb_transactions":  [],
    }

    # Inventory orders — col A "Inventory Orders:" header at row 4,
    # items run from row 5 down. Names in col A, amounts in col B.
    for row_idx in range(5, 100):
        name = sh.cell(row=row_idx, column=COL_INV_NAME).value
        amt  = _num(sh.cell(row=row_idx, column=COL_INV_AMT).value)
        if not name:
            break
        out["inventory"].append({"name": str(name).strip(), "amount": amt})

    # Monthly expenses (Misc / Gas) — typically rows 5-6, cols D-E
    for row_idx in range(5, 12):
        lbl = sh.cell(row=row_idx, column=COL_MONTHLY_LBL).value
        amt = _num(sh.cell(row=row_idx, column=COL_MONTHLY_AMT).value)
        if not lbl:
            continue
        out["monthly_expenses"].append({"name": str(lbl).strip(), "amount": amt})

    # Expense categories — col H rows 2+ until "Total" or blank.
    for row_idx in range(2, 60):
        lbl = sh.cell(row=row_idx, column=COL_EXP_LABEL).value
        if not lbl:
            break
        amt = _num(sh.cell(row=row_idx, column=COL_EXP_AMT).value)
        out["expenses"].append({
            "name":   str(lbl).strip().rstrip(":"),
            "amount": amt,
        })

    # QB paste area — different tabs anchor differently:
    #   - Glendale: header @ col K, data @ col K
    #   - Fresno:   header @ col L, data @ col M (1-col offset)
    # So we don't trust the header; we scan row 2 in the right region for
    # the first column whose value looks like a date — that's the QB DATE col.
    qb_date_col = None
    for c in range(10, sh.max_column + 1):
        v = sh.cell(row=2, column=c).value
        if isinstance(v, (_dt.date, _dt.datetime)):
            qb_date_col = c
            break
    if qb_date_col is not None:
        for row_idx in range(2, 1000):
            date_cell = sh.cell(row=row_idx, column=qb_date_col).value
            if not date_cell:
                break
            out["qb_transactions"].append({
                "date":     date_cell,
                "from_to":  sh.cell(row=row_idx, column=qb_date_col + 1).value,
                "category": sh.cell(row=row_idx, column=qb_date_col + 2).value,
                "memo":     sh.cell(row=row_idx, column=qb_date_col + 3).value,
                "type":     sh.cell(row=row_idx, column=qb_date_col + 4).value,
                "amount":   _num(sh.cell(row=row_idx, column=qb_date_col + 5).value),
                "tags":     sh.cell(row=row_idx, column=qb_date_col + 6).value,
            })

    # Is this event tab "complete enough" to summarize?
    out["is_complete"] = bool(
        out["net_sales"]
        and out["profit"] is not None
        and out["qb_transactions"]
    )
    return out


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _num(v):
    """Coerce to float; returns None on empty/non-numeric."""
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _int(v):
    n = _num(v)
    return int(n) if n is not None else None


# ---------------------------------------------------------------------------
# CLI for ad-hoc inspection
#
#   python3 scripts/lauren_pnl.py            # list tabs
#   python3 scripts/lauren_pnl.py Nashville  # dump that tab
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import sys, json

    wb = fetch_workbook()
    tabs = list_event_tabs(wb)

    if len(sys.argv) < 2:
        print(f"Loaded P&L Sheet — {len(tabs)} tabs:")
        for t in tabs:
            d = get_event_data(wb, t)
            ns = f"${d['net_sales']:,.0f}" if d['net_sales'] else "—"
            pr = f"${d['profit']:,.0f}" if d['profit'] is not None else "—"
            qb = len(d['qb_transactions'])
            done = "✓" if d['is_complete'] else " "
            print(f"  [{done}]  {t:14s}  sales={ns:>10s}  profit={pr:>10s}  qb_rows={qb}")
        raise SystemExit(0)

    tab = sys.argv[1]
    data = get_event_data(wb, tab)
    # Stringify dates for JSON output
    def _ser(o):
        if isinstance(o, (_dt.date, _dt.datetime)):
            return o.isoformat()
        raise TypeError
    print(json.dumps(data, indent=2, default=_ser, ensure_ascii=False))
