#!/usr/bin/env python3
"""pnl_from_sheet.py — backfill per-event P&L JSON from Lauren's manual Google Sheet.

For historical 2026 events (done by hand before the API pipeline), read each tab of
the master P&L sheet and emit the same event_pnl JSON shape the /pnl/ dashboard renders,
so every past event gets a populated summary page. Sheet is the source of truth here.

Verified layout (per tab): C2 net, D2 tx, E2 avg, F2 profit, A2 date, B2 city.
Inventory in A5.., amounts in B (incl a 'Mystery Box' row). Monthly Misc/Gas in D/E.
Expense categories in H (label) / I (amount, negative): staff names, LYFT, Dinner/Meals,
Extra Payment Bonus, Accommodation, TikTok/Facebook Marketing, Shrink/Gas Truck, Car
rental, Warehouse Worker, Flights, Total Shipping inventory (Driver), Venue Rent.
Validated: Glendale lines sum to its sheet profit ($293.86).
"""
from __future__ import annotations
import io, json, sys, urllib.request, datetime as dt
from pathlib import Path
import openpyxl

SHEET_ID = "1mUn3foJhkJCje9LmJVErxKfFwSWJDgcFaZkg_7oWCM4"
URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"

INV_HEADERS = {"inventory orders:", "amount:", "monthly expensses", "total inventory:", ""}


def fetch_wb():
    req = urllib.request.Request(URL, headers={"User-Agent": "Mozilla/5.0"})
    data = urllib.request.urlopen(req, timeout=60).read()
    return openpyxl.load_workbook(io.BytesIO(data), data_only=True)


def _num(v):
    try: return float(v)
    except (TypeError, ValueError): return None


def classify_h(label):
    l = (label or "").lower().strip()
    if "lyft" in l: return "lyft"
    if "dinner" in l or "meal" in l: return "meals"
    if "tiktok" in l: return "marketing_tiktok"
    if "facebook" in l: return "marketing_meta"
    if any(k in l for k in ("accommodation", "hotel", "bnb")): return "travel"
    if "car rental" in l: return "travel"
    if "flight" in l: return "travel"
    if "venue" in l: return "venue"
    if "shipping" in l: return "shipping"
    if "total inventory" in l: return "_inv_total"
    if l == "total": return "_grand"
    if "expens" in l: return None
    if "dollar tree" in l: return "other"
    if any(k in l for k in ("shrink", "gas truck", "car workers", "truck")): return "other"
    if "warehouse" in l: return "staff"
    if "bonus" in l: return "staff"
    return "staff"  # default: a person's name


def parse_tab(ws):
    net = _num(ws["C2"].value); tx = _num(ws["D2"].value)
    avg = _num(ws["E2"].value); profit = _num(ws["F2"].value)
    date = ws["A2"].value
    if isinstance(date, dt.datetime): date = date.date().isoformat()
    city = ws["B2"].value

    # inventory (A5..A25 / B)
    inv_lines = []; mystery = None
    for r in range(5, 30):
        name = ws.cell(row=r, column=1).value
        if name is None: continue
        if str(name).lower().strip() in INV_HEADERS: continue
        amt = _num(ws.cell(row=r, column=2).value)
        if "mystery" in str(name).lower():
            mystery = abs(amt) if amt is not None else None
            continue
        if amt is None: continue
        inv_lines.append({"supplier": str(name), "invoiced": round(abs(amt), 2),
                          "shipping": 0.0, "status": "invoiced"})

    # monthly Misc/Gas (D/E)
    monthly = 0.0
    for r in range(5, 8):
        lbl = ws.cell(row=r, column=4).value
        amt = _num(ws.cell(row=r, column=5).value)
        if lbl and amt is not None:
            monthly += abs(amt)

    # H/I categories
    buckets = {"staff": 0.0, "lyft": 0.0, "meals": 0.0, "marketing_meta": 0.0,
               "marketing_tiktok": 0.0, "travel": 0.0, "venue": 0.0, "shipping": 0.0, "other": 0.0}
    staff_lines = []
    for r in range(2, 24):
        lbl = ws.cell(row=r, column=8).value
        amt = _num(ws.cell(row=r, column=9).value)
        if lbl is None: continue
        cat = classify_h(str(lbl))
        if cat in (None, "_inv_total", "_grand"): continue
        a = abs(amt) if amt is not None else 0.0
        if cat == "staff" and a > 0:
            staff_lines.append({"name": str(lbl), "amount": round(a, 2)})
        if a:
            buckets[cat] = buckets.get(cat, 0.0) + a
    buckets["other"] += monthly

    grand_total = _num(ws.cell(row=21, column=9).value)  # I21 "Total" of expense categories
    inv_total = round(sum(x["invoiced"] for x in inv_lines), 2)
    return {"net": net, "tx": tx, "avg": avg, "profit": profit, "date": date, "city": city,
            "grand_total": grand_total,
            "inventory_lines": inv_lines, "inventory_total": inv_total, "mystery_box": mystery,
            "buckets": buckets, "staff_lines": staff_lines}


def to_event_pnl(parsed, evkey, state=None):
    b = parsed["buckets"]
    def line(amt, ok=True):
        return {"amount": (round(amt, 2) if amt is not None else None),
                "source": "google_sheet (היסטורי)", "status": "ok" if ok else "n/a", "note": ""}
    expenses = {
        "inventory": line(parsed["inventory_total"]),
        "shipping": line(b.get("shipping", 0.0)),
        "mystery_box": line(parsed["mystery_box"] if parsed["mystery_box"] is not None else 0.0),
        "staff": line(b.get("staff", 0.0)),
        "meals": line(b.get("meals", 0.0)),
        "other": line(b.get("other", 0.0)),
        "marketing_meta": line(b.get("marketing_meta", 0.0)),
        "marketing_tiktok": line(b.get("marketing_tiktok", 0.0)),
        "travel": line(b.get("travel", 0.0)),
        "venue": line(b.get("venue", 0.0)),
        "uline": line(None, ok=False),
        "lyft": line(b.get("lyft", 0.0)),
    }
    known = [v["amount"] for v in expenses.values() if v["amount"] is not None]
    total_exp = round(sum(known), 2)
    net = parsed["net"]
    # The H/I expense section (staff/marketing/travel/venue) is filled only when the
    # grand total (I21) is non-zero. If empty, the sheet is INCOMPLETE for this event.
    # Incomplete = the H/I expense section (staff/marketing/venue) was never filled.
    incomplete = (b.get("staff",0)==0 and b.get("venue",0)==0
                  and b.get("marketing_meta",0)==0 and b.get("marketing_tiktok",0)==0)
    profit = round(net - total_exp, 2) if net is not None else parsed["profit"]
    margin = round(profit / net, 4) if (profit is not None and net) else None
    return {
        "evkey": evkey, "historical": True,
        "event": {"city": parsed["city"], "state": state,
                  "start_date": parsed["date"], "end_date": parsed["date"]},
        "revenue": {"net_sales": net, "gross_sales": None, "tax": None,
                    "transactions": int(parsed["tx"]) if parsed["tx"] else None,
                    "avg_ticket": parsed["avg"], "octopos_cash": None,
                    "source": "google_sheet", "status": "historical"},
        "expenses": expenses, "total_known_expenses": total_exp,
        "profit_preliminary": profit, "margin": margin, "preliminary": incomplete,
        "pending_or_missing": (["צוות/שיווק/נסיעות/מקום (לא הוזנו בגיליון)"] if incomplete else []),
        "manual_overrides": {}, "incomplete": incomplete,
        "sheet_profit": parsed["profit"],
        "cash_check": None, "top_products": [],
        "detail": {"payment_breakdown": {}, "inventory_lines": parsed["inventory_lines"],
                   "staff_lines": parsed["staff_lines"], "manager_expense_lines": [],
                   "manager_name": None, "marketing": {"meta": b.get("marketing_meta"), "tiktok": b.get("marketing_tiktok")},
                   "mystery_box": {"name": "Mystery Box", "units": None, "unit_cost": None,
                                   "cost": parsed["mystery_box"]}},
        "warnings": ([f"⚠ נתונים חלקיים בגיליון — הוזנו רק מלאי + הוצאות חודשיות. חסרות הוצאות צוות/שיווק/נסיעות/מקום, אז הרווח אינו סופי."]
                     if incomplete else
                     ([] if abs((profit or 0) - (parsed["profit"] or 0)) < 5 else
                      [f"רווח מחושב ${profit:,.0f} מול הרשום בגיליון ${parsed['profit']:,.0f} — לבדוק"])),
    }


if __name__ == "__main__":
    wb = fetch_wb()
    p = parse_tab(wb["Glendale"])
    j = to_event_pnl(p, "glendale-2026-01-09", state="AZ")
    print("Glendale: net", p["net"], "| computed profit", j["profit_preliminary"],
          "| sheet profit", p["profit"], "| match:", abs(j["profit_preliminary"]-p["profit"])<5)
    print("  expenses:", {k: v["amount"] for k, v in j["expenses"].items()})
    print("  inv lines:", len(p["inventory_lines"]), "| staff:", len(p["staff_lines"]), "| mystery:", p["mystery_box"])
